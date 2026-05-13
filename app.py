import os
import json
import re
import tempfile
import hmac
import logging
import time
import anthropic
import mammoth
import fitz  # PyMuPDF
from flask import Flask, request, render_template, send_file, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
# 15 MB cap — large enough for typical municipal codes, small enough to survive
# Render's free tier 512 MB worker. Bigger docs should be pasted as text.
app.config['MAX_CONTENT_LENGTH'] = 15 * 1024 * 1024

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
APP_PASSWORD = os.environ.get("APP_PASSWORD")  # If unset, app refuses to start /compare requests.

SYSTEM_PROMPT = """You are a municipal code attorney's assistant. You are comparing an OLD version and a NEW version of a municipal code document to produce a professional redline change log.

For every difference you find between the two versions, classify it as one of:
- ADDED — text present in new version but not old
- DELETED — text present in old version but not new
- AMENDED — section exists in both but wording has changed
- RENUMBERED — section moved to a different number (note both old and new numbers)
- PLACEHOLDER — section exists in new version but content is blank/TBD

For each change, identify:
1. The section number (e.g. § 5-404) — if unclear write "Section number not clear"
2. The section title/heading
3. The change type (ADDED / DELETED / AMENDED / RENUMBERED / PLACEHOLDER)
4. The old text (verbatim, or "N/A" if new section)
5. The new text (verbatim, or "N/A" if deleted section)
6. A brief plain-English note explaining the change (1–2 sentences max), suitable for a city council memo

IMPORTANT RULES:
- Quote old and new text EXACTLY as it appears — do not paraphrase, do not normalize whitespace
- For most sections, quote the full text. Only truncate if a section runs longer than ~500 words; in that case quote the first ~500 words verbatim and append "[truncated — full text in original document]"
- For DELETED sections, ALWAYS quote the full text — an attorney needs to see exactly what is being removed regardless of length
- Do not make legal judgments or recommendations
- Do not invent changes — only report what you can directly observe
- If two sections look similar but wording differs even slightly, mark as AMENDED
- Preserve section order in your output

Respond ONLY with valid JSON in exactly this structure — no preamble, no markdown fences:
{
  "summary": "One paragraph plain-English summary of the overall scope of changes suitable for a council memo cover page.",
  "municipality": "City/town name if identifiable from the documents, else 'Not identified'",
  "total_changes": 0,
  "added": 0,
  "deleted": 0,
  "amended": 0,
  "renumbered": 0,
  "placeholder": 0,
  "changes": [
    {
      "section": "§ X-XXX",
      "title": "Section heading",
      "change_type": "AMENDED",
      "old_text": "Exact verbatim old text",
      "new_text": "Exact verbatim new text",
      "note": "Plain-English explanation for council memo"
    }
  ]
}"""


ALLOWED_EXTENSIONS = {"docx", "pdf", "txt"}


def extract_text(file_path, filename):
    """Extract plain text from .docx, .pdf, or .txt file. Raises ValueError for anything else."""
    if "." not in filename:
        raise ValueError(f"File '{filename}' has no extension. Accepted: .docx, .pdf, .txt")
    ext = filename.lower().rsplit('.', 1)[-1]
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"File type '.{ext}' is not supported. Accepted: .docx, .pdf, .txt. "
            f"For .doc, .rtf, or .odt files, save as .docx or paste the text instead."
        )
    try:
        if ext == 'docx':
            with open(file_path, 'rb') as f:
                result = mammoth.extract_raw_text(f)
            return result.value
        elif ext == 'pdf':
            with fitz.open(file_path) as doc:
                return "\n".join(page.get_text() for page in doc)
        else:  # txt
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                return f.read()
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(
            f"Could not read '{filename}' — the file may be corrupt or password-protected. "
            f"({type(e).__name__})"
        ) from e


def truncate(text, max_chars=180000):
    """Returns (truncated_text, was_truncated, original_length)."""
    if len(text) > max_chars:
        return text[:max_chars] + "\n\n[Document truncated due to length]", True, len(text)
    return text, False, len(text)


def run_comparison(old_text, new_text):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    old_truncated_text, old_was_truncated, old_orig_len = truncate(old_text)
    new_truncated_text, new_was_truncated, new_orig_len = truncate(new_text)

    user_message = {
        "role": "user",
        "content": (
            f"Compare these two versions of a municipal code and return the change log JSON.\n\n"
            f"=== OLD VERSION ===\n{old_truncated_text}\n\n"
            f"=== NEW VERSION ===\n{new_truncated_text}"
        ),
    }

    # Retry transient API failures (529 overloaded, 500/503, rate limits).
    # Each attempt also costs real money, so cap at 3.
    max_attempts = 3
    last_error = None
    message = None
    for attempt in range(1, max_attempts + 1):
        try:
            logger.info("Anthropic API call attempt %d/%d", attempt, max_attempts)
            message = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=16000,
                system=SYSTEM_PROMPT,
                messages=[user_message],
            )
            break
        except (anthropic.APIStatusError, anthropic.APIConnectionError, anthropic.RateLimitError) as e:
            last_error = e
            # Don't retry on 4xx auth/validation errors — they won't fix themselves.
            status = getattr(e, "status_code", None)
            if status is not None and 400 <= status < 500 and status not in (408, 429):
                logger.error("Anthropic API error (non-retryable): %s", e)
                raise ValueError(
                    f"The AI service rejected the request ({status}). "
                    f"Check that the API key is valid and has available credit."
                ) from e
            if attempt == max_attempts:
                logger.error("Anthropic API failed after %d attempts: %s", max_attempts, e)
                raise ValueError(
                    "The AI service is currently unavailable after multiple retries. "
                    "Please try again in a few minutes."
                ) from e
            backoff = 2 ** (attempt - 1)  # 1s, 2s
            logger.warning("Transient API error (attempt %d), retrying in %ds: %s",
                           attempt, backoff, e)
            time.sleep(backoff)

    if message is None:
        # Defensive — loop should always set message or raise.
        raise ValueError("The AI service is currently unavailable. Please try again later.") from last_error

    output_truncated = message.stop_reason == "max_tokens"
    raw = message.content[0].text.strip()
    raw = re.sub(r'^```json\s*', '', raw)
    raw = re.sub(r'\s*```$', '', raw)

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        logger.error("Could not parse AI JSON response. stop_reason=%s, raw[:200]=%r",
                     message.stop_reason, raw[:200])
        if output_truncated:
            raise ValueError(
                "The AI response was cut off before finishing — the documents likely contain "
                "too many changes to fit in one response. Try comparing smaller sections."
            ) from e
        raise ValueError(
            "The AI returned a malformed response. This is usually transient — please try again."
        ) from e

    # Surface truncation warnings to the Excel builder
    warnings = []
    if old_was_truncated:
        warnings.append(
            f"⚠ OLD version was truncated: only the first ~{180000:,} characters "
            f"of {old_orig_len:,} were compared. Changes past that point are NOT in this report."
        )
    if new_was_truncated:
        warnings.append(
            f"⚠ NEW version was truncated: only the first ~{180000:,} characters "
            f"of {new_orig_len:,} were compared. Changes past that point are NOT in this report."
        )
    if output_truncated:
        warnings.append(
            "⚠ AI response hit the output length limit — the change list may be incomplete. "
            "Consider comparing smaller sections separately."
        )
    data["_warnings"] = warnings
    logger.info("Comparison complete: %d changes, %d warnings",
                len(data.get("changes", [])), len(warnings))
    return data


# ── Excel colors ──────────────────────────────────────────────────────────────
RED_FILL   = PatternFill("solid", start_color="FFE0E0")   # light red bg
GREEN_FILL = PatternFill("solid", start_color="E0F0E0")   # light green bg
AMBER_FILL = PatternFill("solid", start_color="FFF3CD")   # amber bg
BLUE_FILL  = PatternFill("solid", start_color="DDEEFF")   # blue bg
GREY_FILL  = PatternFill("solid", start_color="F2F2F2")   # grey bg
HEADER_FILL= PatternFill("solid", start_color="1F497D")   # navy header

RED_FONT   = Font(color="C00000", bold=False, name="Arial", size=10)
GREEN_FONT = Font(color="375623", bold=False, name="Arial", size=10)
BLACK_FONT = Font(color="000000", name="Arial", size=10)
WHITE_BOLD = Font(color="FFFFFF", bold=True, name="Arial", size=10)
TITLE_FONT = Font(color="1F497D", bold=True, name="Arial", size=13)
BOLD_FONT  = Font(bold=True, name="Arial", size=10)

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

TYPE_CONFIG = {
    "ADDED":       {"fill": GREEN_FILL, "font": GREEN_FONT, "label": "ADDED ✚"},
    "DELETED":     {"fill": RED_FILL,   "font": RED_FONT,   "label": "DELETED ✖"},
    "AMENDED":     {"fill": AMBER_FILL, "font": BOLD_FONT,  "label": "AMENDED ⟳"},
    "RENUMBERED":  {"fill": BLUE_FILL,  "font": BLACK_FONT, "label": "RENUMBERED ↕"},
    "PLACEHOLDER": {"fill": GREY_FILL,  "font": BLACK_FONT, "label": "PLACEHOLDER ○"},
}


def write_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if border:    cell.border = border
    if alignment: cell.alignment = alignment
    else:         cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def build_xlsx(data, output_path):
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18

    write_cell(ws1, 1, 1, "MUNICIPAL CODE REVISION — CHANGE LOG", font=TITLE_FONT)
    ws1.merge_cells('A1:B1')

    municipality = data.get("municipality", "Not identified")
    write_cell(ws1, 2, 1, f"Municipality: {municipality}", font=BOLD_FONT)
    ws1.merge_cells('A2:B2')

    write_cell(ws1, 3, 1, "AI-Assisted Redline | Not a substitute for attorney review",
               font=Font(italic=True, color="7F7F7F", name="Arial", size=9))
    ws1.merge_cells('A3:B3')

    # Render any truncation warnings prominently before the stats block.
    warnings = data.get("_warnings", [])
    next_row = 4
    if warnings:
        warn_fill = PatternFill("solid", start_color="FFE0E0")
        warn_font = Font(color="C00000", bold=True, name="Arial", size=10)
        write_cell(ws1, next_row, 1, "⚠ INCOMPLETE COMPARISON — READ BEFORE USING",
                   font=warn_font, fill=warn_fill, border=THIN_BORDER)
        ws1.merge_cells(f'A{next_row}:B{next_row}')
        next_row += 1
        for warning in warnings:
            write_cell(ws1, next_row, 1, warning,
                       font=Font(color="C00000", name="Arial", size=10),
                       fill=warn_fill, border=THIN_BORDER,
                       alignment=Alignment(wrap_text=True, vertical="top"))
            ws1.merge_cells(f'A{next_row}:B{next_row}')
            ws1.row_dimensions[next_row].height = 45
            next_row += 1

    ws1.row_dimensions[next_row].height = 8
    stats_start = next_row + 1

    # Stats block
    stats = [
        ("Total Changes", data.get("total_changes", 0), None),
        ("Added",         data.get("added", 0),         GREEN_FILL),
        ("Deleted",       data.get("deleted", 0),        RED_FILL),
        ("Amended",       data.get("amended", 0),        AMBER_FILL),
        ("Renumbered",    data.get("renumbered", 0),     BLUE_FILL),
        ("Placeholder",   data.get("placeholder", 0),    GREY_FILL),
    ]
    for i, (label, val, fill) in enumerate(stats):
        row = stats_start + i
        write_cell(ws1, row, 1, label, font=BOLD_FONT, fill=fill, border=THIN_BORDER)
        write_cell(ws1, row, 2, val,   font=BLACK_FONT, fill=fill, border=THIN_BORDER,
                   alignment=Alignment(horizontal="center", vertical="top"))

    spacer_row = stats_start + len(stats)
    ws1.row_dimensions[spacer_row].height = 8

    summary_header_row = spacer_row + 1
    summary_body_row = summary_header_row + 1
    write_cell(ws1, summary_header_row, 1, "SUMMARY OF CHANGES", font=BOLD_FONT)
    ws1.merge_cells(f'A{summary_header_row}:B{summary_header_row}')
    write_cell(ws1, summary_body_row, 1, data.get("summary", ""),
               font=BLACK_FONT, border=THIN_BORDER,
               alignment=Alignment(wrap_text=True, vertical="top"))
    ws1.merge_cells(f'A{summary_body_row}:B{summary_body_row}')
    ws1.row_dimensions[summary_body_row].height = 90

    # Color key
    key_spacer = summary_body_row + 1
    ws1.row_dimensions[key_spacer].height = 8
    key_header = key_spacer + 1
    write_cell(ws1, key_header, 1, "COLOR KEY", font=BOLD_FONT)
    ws1.merge_cells(f'A{key_header}:B{key_header}')
    key_rows = [
        ("Green text / fill = ADDED",       GREEN_FILL, GREEN_FONT),
        ("Red text / fill = DELETED",        RED_FILL,   RED_FONT),
        ("Amber fill = AMENDED",             AMBER_FILL, BOLD_FONT),
        ("Blue fill = RENUMBERED",           BLUE_FILL,  BLACK_FONT),
        ("Grey fill = PLACEHOLDER",          GREY_FILL,  BLACK_FONT),
    ]
    for i, (label, fill, font) in enumerate(key_rows):
        row = key_header + 1 + i
        write_cell(ws1, row, 1, label, font=font, fill=fill, border=THIN_BORDER)
        ws1.merge_cells(f'A{row}:B{row}')

    # ── Sheet 2: Change Log ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Change Log")

    col_widths = [12, 28, 16, 42, 42, 36]
    col_names  = ["Section", "Title", "Change Type", "Old Text", "New Text", "Note for Council"]
    for i, (w, name) in enumerate(zip(col_widths, col_names), start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        cell = write_cell(ws2, 1, i, name, font=WHITE_BOLD, fill=HEADER_FILL,
                          border=THIN_BORDER,
                          alignment=Alignment(horizontal="center", vertical="center",
                                              wrap_text=True))
    ws2.row_dimensions[1].height = 22

    changes = data.get("changes", [])
    for row_idx, change in enumerate(changes, start=2):
        ctype = change.get("change_type", "AMENDED").upper()
        cfg   = TYPE_CONFIG.get(ctype, TYPE_CONFIG["AMENDED"])

        old_text = change.get("old_text", "N/A")
        new_text = change.get("new_text", "N/A")

        # Old text cell: red strikethrough font if DELETED or AMENDED
        old_font = Font(color="C00000", strike=True, name="Arial", size=9) \
            if ctype in ("DELETED", "AMENDED") else Font(name="Arial", size=9, color="555555")

        # New text cell: green bold font if ADDED or AMENDED
        new_font = Font(color="375623", bold=(ctype == "ADDED"), name="Arial", size=9) \
            if ctype in ("ADDED", "AMENDED") else Font(name="Arial", size=9, color="555555")

        values_fonts = [
            (change.get("section", ""),  BLACK_FONT),
            (change.get("title", ""),    BLACK_FONT),
            (cfg["label"],               cfg["font"]),
            (old_text,                   old_font),
            (new_text,                   new_font),
            (change.get("note", ""),     BLACK_FONT),
        ]

        for col_idx, (val, font) in enumerate(values_fonts, start=1):
            write_cell(ws2, row_idx, col_idx, val,
                       font=font, fill=cfg["fill"], border=THIN_BORDER)

        ws2.row_dimensions[row_idx].height = 60

    # Freeze header row
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Deleted Sections Only ────────────────────────────────────────
    ws3 = wb.create_sheet("Deleted Sections")
    for i, (w, name) in enumerate(zip(col_widths, col_names), start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        write_cell(ws3, 1, i, name, font=WHITE_BOLD, fill=HEADER_FILL,
                   border=THIN_BORDER,
                   alignment=Alignment(horizontal="center", vertical="center"))
    ws3.row_dimensions[1].height = 22

    del_row = 2
    deletion_count = 0
    for change in changes:
        if change.get("change_type", "").upper() == "DELETED":
            deletion_count += 1
            cfg = TYPE_CONFIG["DELETED"]
            vals = [
                change.get("section", ""), change.get("title", ""),
                cfg["label"], change.get("old_text", ""), "N/A",
                change.get("note", "")
            ]
            fonts = [BLACK_FONT, BLACK_FONT, RED_FONT,
                     Font(color="C00000", strike=True, name="Arial", size=9),
                     BLACK_FONT, BLACK_FONT]
            for col_idx, (val, font) in enumerate(zip(vals, fonts), start=1):
                write_cell(ws3, del_row, col_idx, val,
                           font=font, fill=RED_FILL, border=THIN_BORDER)
            ws3.row_dimensions[del_row].height = 60
            del_row += 1

    if deletion_count == 0:
        write_cell(ws3, 2, 1,
                   "No deletions were found in this comparison.",
                   font=Font(italic=True, color="7F7F7F", name="Arial", size=10),
                   alignment=Alignment(horizontal="center", vertical="center"))
        ws3.merge_cells('A2:F2')
        ws3.row_dimensions[2].height = 30

    ws3.freeze_panes = "A2"

    wb.save(output_path)


@app.errorhandler(413)
def handle_too_large(e):
    return jsonify({
        "error": "File is too large (15 MB max). For very large documents, "
                 "paste the text directly instead of uploading."
    }), 413


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/compare", methods=["POST"])
def compare():
    start = time.monotonic()
    if not ANTHROPIC_API_KEY:
        logger.error("ANTHROPIC_API_KEY missing")
        return jsonify({"error": "ANTHROPIC_API_KEY not configured on server"}), 500
    if not APP_PASSWORD:
        logger.error("APP_PASSWORD missing")
        return jsonify({"error": "APP_PASSWORD not configured on server"}), 500

    # Accept password from header (preferred) or form field (browser convenience).
    submitted = request.headers.get("X-App-Password") or request.form.get("app_password", "")
    # Constant-time compare to avoid timing attacks on short secrets.
    if not hmac.compare_digest(submitted, APP_PASSWORD):
        logger.warning("Failed auth attempt from %s", request.remote_addr)
        return jsonify({"error": "Invalid or missing password."}), 401

    old_file = request.files.get("old_file")
    new_file = request.files.get("new_file")
    old_text_raw = request.form.get("old_text", "").strip()
    new_text_raw = request.form.get("new_text", "").strip()

    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            # Resolve old text
            if old_file and old_file.filename:
                safe_old = secure_filename(old_file.filename) or "old_upload"
                old_path = os.path.join(tmpdir, "old_" + safe_old)
                old_file.save(old_path)
                old_text = extract_text(old_path, old_file.filename)
            elif old_text_raw:
                old_text = old_text_raw
            else:
                return jsonify({"error": "Please provide the old version (file or paste)."}), 400

            # Resolve new text
            if new_file and new_file.filename:
                safe_new = secure_filename(new_file.filename) or "new_upload"
                new_path = os.path.join(tmpdir, "new_" + safe_new)
                new_file.save(new_path)
                new_text = extract_text(new_path, new_file.filename)
            elif new_text_raw:
                new_text = new_text_raw
            else:
                return jsonify({"error": "Please provide the new version (file or paste)."}), 400
        except ValueError as e:
            logger.info("Extraction error: %s", e)
            return jsonify({"error": str(e)}), 400

        if not old_text.strip() or not new_text.strip():
            return jsonify({"error": "Could not extract text from one or both documents."}), 400

        logger.info("Starting comparison: old=%d chars, new=%d chars",
                    len(old_text), len(new_text))

        try:
            data = run_comparison(old_text, new_text)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        except Exception as e:
            logger.exception("Unexpected error during comparison")
            return jsonify({"error": f"Unexpected server error: {type(e).__name__}"}), 500

        output_path = os.path.join(tmpdir, "change_log.xlsx")
        try:
            build_xlsx(data, output_path)
        except Exception as e:
            logger.exception("Failed to build Excel file")
            return jsonify({"error": f"Failed to build Excel file: {type(e).__name__}"}), 500

        elapsed = time.monotonic() - start
        logger.info("Comparison finished in %.1fs", elapsed)

        return send_file(
            output_path,
            as_attachment=True,
            download_name="Municipal_Code_Change_Log.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    app.run(host="0.0.0.0", port=port)
